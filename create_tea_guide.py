# -*- coding: utf-8 -*-
"""
일본 차 구입 가이드 엑셀 파일 생성 스크립트
후쿠오카(텐진/하카타) 여행자용 - 2026년 4월 8~10일
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# 스타일 정의
# ============================================================
HEADER_FILL = PatternFill("solid", fgColor="2D6A2D")   # 진한 녹색
SUBHEADER_FILL = PatternFill("solid", fgColor="5A9A5A") # 중간 녹색
ROW_ALT_FILL = PatternFill("solid", fgColor="EAF4EA")  # 연한 녹색
OVERVIEW_FILL = PatternFill("solid", fgColor="3A7A3A")
TITLE_FILL = PatternFill("solid", fgColor="1A4A1A")

WHITE_FONT = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=11)
HEADER_FONT = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Malgun Gothic", size=9)
BOLD_FONT = Font(name="Malgun Gothic", bold=True, size=9)
TITLE_FONT = Font(name="Malgun Gothic", bold=True, color="FFFFFF", size=14)

BORDER_THIN = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_header(ws, row, cols_data, fill=HEADER_FILL):
    for col, text in enumerate(cols_data, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER_THIN


def set_row(ws, row, cols_data, alt=False):
    fill = ROW_ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    for col, text in enumerate(cols_data, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = fill
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER_THIN


def auto_width(ws, min_w=12, max_w=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                length = max(len(line) for line in lines)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max(max_len * 1.2, min_w), max_w)


def title_row(ws, text, row=1, ncols=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = CENTER
    ws.row_dimensions[row].height = 32


# ============================================================
# Sheet 1: 일본 차 문화 개요
# ============================================================
ws1 = wb.active
ws1.title = "1_일본차문화개요"

title_row(ws1, "일본 차 문화 개요 & 후쿠오카(야메차) 특색", row=1, ncols=4)

# 섹션별 개요 내용
overview_data = [
    ("섹션", "항목", "내용", "비고"),
    ("일본 차 문화\n기본 이해",
     "센차(煎茶)란?",
     "증기로 쪄서 만드는 일본 대표 녹차.\n수확 후 즉시 가열 처리하여 산화를 막음.\n일본 녹차 생산의 약 60~70% 차지.",
     "초보자에게\n가장 추천"),
    ("일본 차 문화\n기본 이해",
     "센차도(煎茶道)",
     "센차를 마시고 즐기는 전통 의식 문화.\n말차(抹茶)와 달리 일상적으로 즐기는 차 문화.\n도구: 교스(きゅうす) - 옆손잡이 찻주전자",
     ""),
    ("일본 차 문화\n기본 이해",
     "올바른 우리기",
     "물 온도: 70~80°C (끓인 물을 식혀서 사용)\n찻잎량: 4~5g / 200ml\n우리는 시간: 45~90초\n⚠️ 펄펄 끓는 물 사용 시 쓴맛 과다 추출",
     "온도가\n핵심!"),
    ("수확 시기별\n등급",
     "신차·이치반차\n(新茶·一番茶)",
     "4월 말~5월 초 첫 수확. 최고 품질.\n낮은 카페인, 높은 아미노산(테아닌), 낮은 카테킨.\n섬세한 맛, 선명한 녹색, 풍부한 향.",
     "4월 방문 시\n신차 구매 적기!"),
    ("수확 시기별\n등급",
     "니반차(二番茶)",
     "6~7월 두 번째 수확. 중급.\n신차보다 향이 다소 약하고 가격 합리적.",
     ""),
    ("수확 시기별\n등급",
     "삼반차 이후\n(三番茶~)",
     "여름~가을 세 번째 이후 수확. 하급.\n쓴맛 강함. 일상 소비용. 가장 저렴.",
     ""),
    ("후쿠오카\n야메차(八女茶)\n특색",
     "생산 지역",
     "후쿠오카현 남부 야베강·호시노강 유역.\n야메시(八女市), 치쿠고시(筑後市), 히로카와정,\n우키하시(浮羽市), 아사쿠라시(朝倉市)",
     "텐진에서\n약 1시간 거리"),
    ("후쿠오카\n야메차(八女茶)\n특색",
     "역사",
     "1423년 선종 승려 에이린 슈즈이가 명나라에서\n차 씨앗을 가져옴. 에도시대부터 본격 재배.\n현재 일본 전체 교쿠로(玉露) 생산의 약 40~50% 차지.",
     "약 600년\n역사"),
    ("후쿠오카\n야메차(八女茶)\n특색",
     "왜 맛있는가?",
     "내륙성 기후: 낮은 고온 + 밤의 급격한 냉각 + 아침 안개\n→ 우마미 성분(테아닌·L-테아닌) 풍부하게 축적\n결과: 쓴맛 적고, 깊은 단맛 + 강한 감칠맛(우마미)",
     ""),
    ("후쿠오카\n야메차(八女茶)\n특색",
     "타 지역 차와 차이",
     "우지(宇治)·시즈오카(静岡): 밝고 풀 향기 강함\n야메차: 부드럽고 자연스러운 단맛,\n'버터 같은' 또는 '견과류 같은' 고소한 향이 특징",
     ""),
    ("후쿠오카\n야메차(八女茶)\n특색",
     "4월 방문 팁",
     "4월 8~10일은 신차(新茶) 시즌 직전!\n매장에서 전년도 최고급 야메차 또는\n신차 예약 판매를 확인해보세요.\n야메 이치반차는 4월 말~5월 초 수확.",
     "⭐ 여행 타이밍\n최적!"),
    ("규슈 차 산지\n비교",
     "후쿠오카 야메",
     "대표 차: 교쿠로·센차·가부세차\n특징: 일본 교쿠로 생산 1위, 깊은 우마미",
     ""),
    ("규슈 차 산지\n비교",
     "사가 우레시노\n(嬉野)",
     "대표 차: 카마이리차(釜炒り茶)·타마료쿠차\n특징: 둥근 잎 모양, 고소한 볶음 향",
     "텐진에서\n당일치기 가능"),
    ("규슈 차 산지\n비교",
     "미야자키 기리시마\n(霧島)",
     "대표 차: 고급 센차·카마이리차\n특징: 풍부한 일조량, 독특한 카마코(釜香) 향",
     ""),
]

set_header(ws1, 2, ["섹션", "항목", "내용", "비고"])
for i, row_data in enumerate(overview_data[1:], start=3):
    set_row(ws1, i, row_data, alt=(i % 2 == 0))
    ws1.row_dimensions[i].height = 60

ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 55
ws1.column_dimensions["D"].width = 18
ws1.freeze_panes = "A3"


# ============================================================
# Sheet 2: 차 종류별 상세 가이드
# ============================================================
ws2 = wb.create_sheet("2_차종류별상세가이드")

title_row(ws2, "후쿠오카(야메) 잎차(센차) 종류별 상세 가이드", row=1, ncols=9)

headers2 = [
    "일본어 표기", "한글 읽기", "차 종류\n(재배 방식)", "역사 및 산지",
    "맛과 향의 특징", "추천 대상", "가격\n(g당 엔)", "등급",
    "참고 이미지 / 공식 사이트 URL"
]
set_header(ws2, 2, headers2)

tea_data = [
    (
        "煎茶\n(センチャ)",
        "센차",
        "비차광 재배\n(증제)",
        "야메차(八女茶)의 기본.\n야메 지역 후쿠오카현 남부 생산.\n에도시대부터 재배, 600년 역사.",
        "【맛】 신선하고 청량감 있는 풀 향.\n균형 잡힌 단맛·감칠맛·쓴맛.\n【향】 밝고 상큼한 초록빛 향기.\n【색】 맑은 황록색.\n초보자도 부담 없이 즐길 수 있음.",
        "처음 일본 차를\n접하는 분\n일상 음용",
        "10~20엔/g\n(100g당\n1,000~2,000엔)",
        "중급",
        "https://fukuoka-yamecha.jp/green-tea-yame/\nhttps://www.yamechanosato.com/"
    ),
    (
        "深蒸し煎茶\n(フカムシ\nセンチャ)",
        "후카무시\n센차",
        "비차광 재배\n(장시간 증제)",
        "야메 지역 대표 센차 중 하나.\n일반 센차보다 2~3배 길게 증제하여\n진하고 부드러운 맛 완성.",
        "【맛】 진하고 부드러운 맛.\n쓴맛 적고 단맛·감칠맛 강함.\n【향】 풍부하고 깊은 녹차 향.\n【색】 진한 녹색 (잎이 미세하게 분쇄됨).\n초보자에게 특히 추천!",
        "진한 녹차 맛\n선호하는 분\n(가장 추천!)",
        "13~25엔/g\n(100g당\n1,350~2,500엔)",
        "중~상급",
        "https://shop.koganoyamecha.co.jp/\nhttps://yamecha.biz/"
    ),
    (
        "かぶせ茶\n(カブセチャ)",
        "가부세차",
        "부분 차광 재배\n(수확 2주 전\n차광막 설치)",
        "야메 지역 특산.\n수확 2주 전 차광으로 테아닌(감칠맛 성분)\n생산 촉진. 센차와 교쿠로의 중간 등급.",
        "【맛】 센차의 신선함 + 교쿠로의 감칠맛 조화.\n단맛과 우마미(감칠맛) 모두 강함.\n【향】 진한 녹색 향기, 약간 해초향.\n【색】 진한 녹색.",
        "센차와 교쿠로\n사이 등급 원하는 분\n가성비 고급 차",
        "20~35엔/g\n(100g당\n2,000~3,500엔)",
        "중~상급",
        "https://www.irie-chaen.com/\nhttps://chanohado.com/"
    ),
    (
        "玉露\n(ギョクロ)",
        "교쿠로",
        "강 차광 재배\n(수확 20~30일 전\n90% 이상 차광)",
        "야메 지역이 일본 교쿠로 생산 1위.\n2007년 전국 교쿠로 품질 경연대회\n상위 26개 브랜드 전부 야메산.\n최고급 일본 녹차.",
        "【맛】 강력한 우마미(감칠맛) + 자연스러운 단맛.\n쓴맛 거의 없음.\n【향】 버터 같은 부드러운 향, 고소함.\n【색】 짙은 황록색.\n※ 물 온도 50~60°C로 낮춰 우려야 함!",
        "일본 차 마니아\n최고급 선물 구매",
        "30~80엔/g\n(100g당\n3,000~8,000엔)",
        "고급~최고급",
        "https://www.hoshitea.com/\nhttps://www.yame.co.jp/\nhttps://www.o-cha.com/en/gyokuro/yame-gyokuro.html"
    ),
    (
        "星野玉露\n(ホシノ\nギョクロ)",
        "호시노\n교쿠로",
        "강 차광 재배\n(야메 호시노마을\n특산)",
        "야메시 호시노마을(星野村) 생산.\n호시노강 근처 미세기후 + 아침저녁 안개로\n천연 차광 효과. 2014년 일본의 아름다운\n마을 선정 지역.",
        "【맛】 교쿠로 중 최고급 우마미 + 단맛.\n해양성 감칠맛, 녹아드는 부드러운 맛.\n【향】 버터 같은 고급스러운 향.\n【색】 짙은 황금빛 녹색.",
        "최고급 교쿠로\n선물용 추천",
        "50~120엔/g\n(80g당\n4,000~9,600엔)",
        "최고급",
        "https://www.hoshitea.com/\nhttps://www.yoshien.com/en/gyokuro-hoshino.html"
    ),
    (
        "玉緑茶・ぐり茶\n(タマリョク\nチャ)",
        "타마료쿠차\n(구리차)",
        "비차광 재배\n(증제 or 솥 볶기\n+ 특수 마무리)",
        "규슈 특유의 차 종류.\n마무리 과정에서 비틀기를 하지 않아\n둥그스름한 곱슬 모양의 찻잎.\n사가 우레시노와 후쿠오카에서 생산.",
        "【맛】 강한 감칠맛 + 우마미.\n우릴수록 변화하는 층층이 깊은 맛.\n【향】 진하고 풍부한 녹차 향.\n【색】 진한 녹색.\n찻잎이 서서히 펼쳐지는 시각적 즐거움.",
        "특이한 모양\n원하는 분\n규슈 특산 선물",
        "15~30엔/g\n(100g당\n1,500~3,000엔)",
        "중급",
        "https://www.yamechanosato.com/"
    ),
    (
        "釜炒り茶\n(カマイリチャ)",
        "카마이리차",
        "비차광 재배\n(증제 아닌\n솥 볶기 방식)",
        "일본 전통 제다법 중 하나.\n사가현 우레시노(嬉野)가 대표 산지.\n전체 일본 차 중 약 5%만 생산되는 희귀 차.\n중국 제다법의 영향을 받음.",
        "【맛】 약한 쓴맛 + 상큼하고 청량한 맛.\n아몬드 같은 고소한 견과류 향.\n【향】 카마코(釜香): 솥 볶음 특유의\n독특한 고소한 볶음 향기.\n【색】 밝은 황금빛.",
        "독특한 향 원하는 분\n중국 녹차 좋아하는 분",
        "15~35엔/g\n(100g당\n1,500~3,500엔)",
        "중급",
        "https://kimikuratea.com/products/kamairicha-green-tea-ureshino-saga\nhttps://en.wikipedia.org/wiki/Kamairicha"
    ),
]

for i, row_data in enumerate(tea_data, start=3):
    set_row(ws2, i, row_data, alt=(i % 2 == 0))
    ws2.row_dimensions[i].height = 90

ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 30
ws2.column_dimensions["E"].width = 40
ws2.column_dimensions["F"].width = 16
ws2.column_dimensions["G"].width = 16
ws2.column_dimensions["H"].width = 10
ws2.column_dimensions["I"].width = 45
ws2.freeze_panes = "A3"


# ============================================================
# Sheet 3: 구매 가능 매장 정보
# ============================================================
ws3 = wb.create_sheet("3_구매매장정보")

title_row(ws3, "후쿠오카(텐진·하카타) 일본 차 구매 매장 정보", row=1, ncols=9)

headers3 = [
    "매장명\n(일본어)", "매장명\n(한글)", "주소", "텐진/하카타역\n에서 접근",
    "영업시간", "휴무일", "주요 취급 차", "시음\n가능", "특징 및 메모"
]
set_header(ws3, 2, headers3)

store_data = [
    (
        "茶舗ふりゅう\n(天神薬院店)",
        "차호 후류우\n(텐진 야쿠인점)",
        "〒810-0005\n福岡市中央区清川1丁目6-9 2F",
        "텐진역에서 도보 10분\n와타나베도리역에서 도보 3분\n서철 야쿠인역에서 도보 7분",
        "11:00~18:00\n(월요일 11:00~17:00)",
        "화·수요일 및\n부정기 휴무",
        "야메 교쿠로(玉露)\n가부세차\n고급 야메차 전반",
        "⭕ 가능\n(카운터 시음\n비교 가능)",
        "야메 교쿠로 전문 매장.\n장인 도자기 다구도 함께 판매.\n소규모 고급 전문점.\nEmail: contact@chahofuryu.com"
    ),
    (
        "XINFUCHA LAB\n(新福茶ラボ)",
        "신푸차 랩",
        "福岡市中央区天神2-2-43\nソラリアプラザ B2F",
        "텐진역에서 바로\n소라리아 플라자 지하 2층",
        "11:00~20:00",
        "시설 휴관일",
        "규슈산 오리지널 블렌드 센차\n후쿠오카 야메 센차\n야쿠시마 홍차\n미야자키 카마이리차",
        "⭕ 가능\n(매장 내 시음)",
        "2024년 12월 오픈 신규 매장.\n규슈 각지의 차를 블렌드한 오리지널 제품 특화.\n쇼핑 중 들르기 편리한 위치.\nTEL: 092-406-7655"
    ),
    (
        "光安青霞園茶舗\n(こうやすせいかえん)",
        "광안 청하원\n차포",
        "福岡市博多区\n(하카타 시내)",
        "하카타역에서\n도보권 내",
        "요문의",
        "요문의",
        "야메차(八女茶) 전반\n센차·교쿠로",
        "⭕ 가능\n(노포 시음)",
        "창업 300년 이상의 노포(老舗).\n야메차 전통 판매점.\n역사적인 분위기에서 차 구매 가능.\n하카타 관광 시 들르기 좋음."
    ),
    (
        "大濠テラス\n「八女茶と日本庭園」",
        "대호 테라스\n'야메차와 일본 정원'",
        "福岡市中央区\n大濠公園 南側",
        "텐진역에서\n지하철 오호리공원역\n도보 약 5분",
        "10:00~21:00\n(계절에 따라 변동)",
        "연중무휴\n(시설에 따름)",
        "야메차(八女茶)\n각종 야메차 상품",
        "⭕ 가능\n(카페 형식 시음)",
        "오호리공원 내 위치.\n야메차와 일본 화과자를 함께 즐기는 공간.\n공원 산책 후 들르기 최적의 위치.\n경치 좋은 일본 정원 분위기."
    ),
    (
        "LUPICIA\n(ルピシア)\n福岡天神店",
        "루피시아\n후쿠오카 텐진점",
        "福岡市中央区天神\n(텐진 쇼핑몰 내)",
        "텐진역에서\n도보 5분 이내",
        "쇼핑몰 영업시간에 준함\n통상 10:00~20:00",
        "시설 휴관일",
        "세계 각지의 녹차·홍차·우롱차\n일본 각지 센차\n허브티 등 200종 이상",
        "⭕ 가능\n(계절 시음 제공)",
        "세계 최대급 차 전문 브랜드.\n일본 각지 고급 녹차 포함 200종 이상.\n예쁜 패키지로 선물용으로 인기.\n영어 응대 가능한 경우 있음."
    ),
    (
        "JR博多シティ\nAMU PLAZA 내 차 매장",
        "JR 하카타시티\n아무 플라자 내 차 매장",
        "福岡市博多区博多駅中央街1-1\nJR博多シティ",
        "하카타역 직결\n(역 건물 내)",
        "10:00~21:00",
        "연중무휴\n(일부 매장 제외)",
        "후쿠오카·규슈 특산 차\n야메차 상품 다수\n선물용 포장 차",
        "매장에 따라\n다름",
        "하카타역 직결 쇼핑몰.\n관광 기념품으로 야메차 세트 구매 최적.\nB1~8F의 대형 쇼핑몰.\n출발 전 마지막 쇼핑 장소로 추천."
    ),
]

for i, row_data in enumerate(store_data, start=3):
    set_row(ws3, i, row_data, alt=(i % 2 == 0))
    ws3.row_dimensions[i].height = 85

ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 28
ws3.column_dimensions["D"].width = 22
ws3.column_dimensions["E"].width = 18
ws3.column_dimensions["F"].width = 16
ws3.column_dimensions["G"].width = 25
ws3.column_dimensions["H"].width = 12
ws3.column_dimensions["I"].width = 40
ws3.freeze_panes = "A3"


# ============================================================
# Sheet 4: 시음 가능 매장 상세
# ============================================================
ws4 = wb.create_sheet("4_시음가능매장")

title_row(ws4, "후쿠오카 일본 차 시음(試飮) 가능 매장 상세", row=1, ncols=7)

headers4 = [
    "매장명", "위치 및 접근", "시음 방식",
    "시음 가능 차 종류", "시음 비용", "영업시간", "추천 포인트"
]
set_header(ws4, 2, headers4)

tasting_data = [
    (
        "茶舗ふりゅう\n(차호 후류우)",
        "텐진역에서 도보 10분\n福岡市中央区清川1-6-9 2F\n와타나베도리역 도보 3분",
        "카운터에서\n스태프와 함께\n복수 차 비교 시음\n(테이크아웃 가능)",
        "야메 교쿠로(玉露)\n가부세차(かぶせ茶)\n고급 야메 센차\n(시즌별 변동)",
        "무료\n(구매 전 시음 형식)",
        "11:00~18:00\n(월 ~17:00)\n화·수 휴무",
        "전문 스태프가 차 설명해줌.\n교쿠로와 가부세차를 비교할 수 있는\n드문 기회. 소규모 고급 전문점."
    ),
    (
        "XINFUCHA LAB\n(신푸차 랩)",
        "텐진역 직결\n소라리아 플라자 B2F\n092-406-7655",
        "매장 내 시음 코너에서\n오리지널 블렌드 차 시음\n(스탠딩 또는 착석)",
        "규슈 오리지널 블렌드 센차\n후쿠오카 야메 센차\n미야자키 카마이리차\n야쿠시마 홍차",
        "무료~소액\n(상품에 따라)",
        "11:00~20:00\n시설 휴관일 제외",
        "2024년 12월 오픈 신규 매장.\n텐진 쇼핑 중 부담 없이 들를 수 있음.\n규슈 각지의 다양한 차 한번에 비교 가능."
    ),
    (
        "大濠テラス\n'야메차와 일본 정원'",
        "텐진역에서 지하철\n오호리공원역 도보 5분\n오호리공원 남측",
        "카페 형식\n(테이블에서 음식과 함께\n차를 주문해서 시음)",
        "야메차 각종\n(센차·교쿠로·가부세차)\n계절 한정 차",
        "300~800엔\n(음료 주문 형식)",
        "10:00~21:00\n(계절 변동)",
        "일본 정원+오호리공원 경치 감상하며\n야메차를 천천히 즐길 수 있음.\n화과자와 함께 즐기는 정통 일본 차 체험."
    ),
    (
        "茶の文化館\n(차의 문화관)",
        "야메시 호시노마을\n(텐진에서 차로 약 1시간)\n호시노 마을 내",
        "체험 프로그램 참가\n+ 다양한 차 자유 시음",
        "호시노 교쿠로(星野玉露)\n야메차 전 종류\n말차(抹茶) 체험 포함",
        "입장 무료\n말차 갈기 체험: 500엔\n(화과자 포함)\n말차 점차 체험: +300엔\n손으로 비빈 녹차 만들기: 1,000엔",
        "10:00~17:00\n화요일 휴관\n(공휴일·5월·여름방학·정월은 개관)",
        "⭐강력 추천!\n야메 현지에서 최고급 호시노 교쿠로 시음 가능.\n차 만들기 체험까지 가능.\n당일치기 야메 투어 계획 시 필수 방문지."
    ),
    (
        "LUPICIA\n후쿠오카 텐진점",
        "텐진역에서 도보 5분\n텐진 쇼핑몰 내",
        "계절별 시음 코너에서\n무료 시음 제공\n(스태프 안내)",
        "계절 추천 일본 녹차\n세계 각지 홍차·우롱차\n신상품 중심 시음",
        "무료",
        "쇼핑몰 영업시간 준함\n통상 10:00~20:00",
        "200종 이상의 차 중 선택 가능.\n예쁜 패키지로 선물 구매에 최적.\n영어 안내 가능."
    ),
]

for i, row_data in enumerate(tasting_data, start=3):
    set_row(ws4, i, row_data, alt=(i % 2 == 0))
    ws4.row_dimensions[i].height = 90

ws4.column_dimensions["A"].width = 18
ws4.column_dimensions["B"].width = 25
ws4.column_dimensions["C"].width = 22
ws4.column_dimensions["D"].width = 28
ws4.column_dimensions["E"].width = 18
ws4.column_dimensions["F"].width = 16
ws4.column_dimensions["G"].width = 42
ws4.freeze_panes = "A3"


# ============================================================
# Sheet 5: 여행 팁 & 빠른 참고
# ============================================================
ws5 = wb.create_sheet("5_여행팁_빠른참고")

title_row(ws5, "4월 8~10일 후쿠오카 차 여행 빠른 참고 가이드", row=1, ncols=4)

tip_headers = ["카테고리", "항목", "내용", "메모"]
set_header(ws5, 2, tip_headers)

tip_data = [
    ("구매 추천 순위", "1순위 (최고급 선물)", "호시노 교쿠로(星野玉露)\n야메 교쿠로(八女玉露)",
     "80g 3,000~8,000엔\n고급 선물 포장 가능"),
    ("구매 추천 순위", "2순위 (가성비 최고)", "후카무시 센차(深蒸し煎茶)\n또는 가부세차(かぶせ茶)",
     "100g 1,350~3,000엔\n마시기 쉽고 맛 보장"),
    ("구매 추천 순위", "3순위 (체험·기념)", "타마료쿠차(玉緑茶) 둥근 잎 모양\n카마이리차(釜炒り茶) 볶음 향",
     "독특한 모양/향으로\n기념품 추천"),
    ("가격 기준", "저렴한 센차", "g당 10~15엔 (100g 1,000~1,500엔)", "일상 음용"),
    ("가격 기준", "중급 센차·가부세차", "g당 20~35엔 (100g 2,000~3,500엔)", "품질과 가격 밸런스 최고"),
    ("가격 기준", "고급 교쿠로", "g당 30~80엔 (100g 3,000~8,000엔)", "선물용"),
    ("가격 기준", "최고급 호시노 교쿠로", "g당 80~120엔 이상 (80g 6,400엔~)", "최고급 선물"),
    ("쇼핑 코스\n(텐진 출발)", "코스 A: 텐진 집중",
     "① XINFUCHA LAB (소라리아 B2F)\n② 茶舗ふりゅう (도보 10분)\n③ 대호 테라스 (지하철 이동)",
     "텐진 내 반나절 코스"),
    ("쇼핑 코스\n(하카타역)", "코스 B: 하카타역",
     "① JR하카타시티 아무 플라자 (차 코너)\n② 광안 청하원 차포 (창업 300년)",
     "역 건물 내 편리"),
    ("쇼핑 코스\n(당일치기)", "코스 C: 야메 당일치기",
     "텐진→(버스·렌터카 1시간)→\n야메시 차의 문화관 → 호시노마을\n→ 현지 차원 직구매",
     "⭐ 4월은 신차 시즌 직전!\n현지 구매 시 가장 신선"),
    ("차 포장·운반", "기내 반입", "밀봉 포장 차는 기내 반입 가능.\n50ml 이상 액체는 위탁 수하물로.",
     "차잎 자체는 문제없음"),
    ("차 포장·운반", "보관 방법", "직사광선·열·습기 피해 보관.\n개봉 후 냉장 또는 서늘한 곳 보관.\n최대 6개월 내 음용 권장.",
     "냉동 보관도 가능"),
    ("언어 팁", "차 주문 시 일본어",
     "「これを試飲できますか？」\n(이거 시음 가능한가요?)\n「おすすめはどれですか？」\n(추천은 어느 것인가요?)\n「贈り物用に包んでください」\n(선물용으로 포장해주세요)",
     "매장 직원에게 사용"),
    ("결제", "결제 수단", "현금(엔화) 선호 매장 多.\nICカード(Suica 등)·신용카드:\n대형 매장은 대부분 가능.\n소규모 전문점은 현금 권장.",
     "현금 준비 권장"),
]

for i, row_data in enumerate(tip_data, start=3):
    set_row(ws5, i, row_data, alt=(i % 2 == 0))
    ws5.row_dimensions[i].height = 55

ws5.column_dimensions["A"].width = 18
ws5.column_dimensions["B"].width = 22
ws5.column_dimensions["C"].width = 50
ws5.column_dimensions["D"].width = 22
ws5.freeze_panes = "A3"


# ============================================================
# 저장
# ============================================================
output_path = "/Users/choongheeseo/antigravity_002/일본차_구입가이드_후쿠오카.xlsx"
wb.save(output_path)
print(f"✅ 엑셀 파일 생성 완료: {output_path}")

# 검증
from openpyxl import load_workbook
wb2 = load_workbook(output_path)
print(f"시트 목록: {wb2.sheetnames}")
for sname in wb2.sheetnames:
    ws = wb2[sname]
    print(f"  [{sname}] 행 수: {ws.max_row}, 열 수: {ws.max_column}")
print("✅ 검증 완료")
